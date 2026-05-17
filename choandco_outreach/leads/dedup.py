import hashlib
import json
import logging
from datetime import datetime

from config import DEDUP_PATH, EXCEL_PATH, DATA_DIR
from leads.models import Lead

logger = logging.getLogger(__name__)

_store = None


def _normalize_legacy_entry(entry: dict) -> dict:
    """Upgrade a legacy 3-field entry to the full sequence-tracking schema."""
    if "sequence_stage" not in entry:
        entry["decision_maker_name"] = entry.get("decision_maker_name", "")
        entry["industry"] = entry.get("industry", "")
        entry["original_subject"] = entry.get("original_subject", "")
        entry["initial_sent_date"] = entry.get("sent_date") or entry.get("initial_sent_date", "")
        entry["sequence_stage"] = "initial"
        entry["followup1_date"] = None
        entry["followup2_date"] = None
        entry["followup3_date"] = None
        # Remove legacy key
        entry.pop("sent_date", None)
    return entry


def _load_store() -> dict:
    """Load the dedup hash store from disk, normalizing legacy entries."""
    global _store
    if _store is not None:
        return _store

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    if DEDUP_PATH.exists():
        try:
            with open(DEDUP_PATH, "r") as f:
                raw = json.load(f)
            # Normalize any legacy entries
            needs_save = False
            for h, entry in raw.items():
                if "sequence_stage" not in entry:
                    _normalize_legacy_entry(entry)
                    needs_save = True
            _store = raw
            if needs_save:
                _save_store()
                logger.info("Migrated legacy dedup entries to expanded schema")
            return _store
        except (json.JSONDecodeError, IOError):
            logger.warning("Could not read dedup store, rebuilding from Excel")

    _store = _rebuild_from_excel()
    _save_store()
    return _store


def _rebuild_from_excel() -> dict:
    """Rebuild the hash store from the Excel outreach log."""
    store = {}
    if not EXCEL_PATH.exists():
        return store

    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_PATH, read_only=True)
        ws = wb.active
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        is_new_schema = header_row and len(header_row) >= 12

        for row in ws.iter_rows(min_row=2, values_only=True):
            if is_new_schema:
                company = str(row[0] or "")
                dm_name = str(row[1] or "")
                email = str(row[3] or "")
                sent_date = str(row[11] or "")
            else:
                company = str(row[0] or "")
                dm_name = ""
                email = str(row[1] or "")
                sent_date = str(row[4] or "") if row[4] else ""

            if company and email:
                h = compute_hash(company, email)
                store[h] = _normalize_legacy_entry({
                    "company": company,
                    "email": email,
                    "decision_maker_name": dm_name,
                    "industry": "",
                    "original_subject": "",
                    "sent_date": sent_date,
                })
        wb.close()
        logger.info("Rebuilt dedup store from Excel: %d entries", len(store))
    except Exception as e:
        logger.error("Failed to rebuild from Excel: %s", e)

    return store


def _save_store() -> None:
    """Persist the hash store to disk."""
    global _store
    if _store is None:
        return
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(DEDUP_PATH, "w") as f:
        json.dump(_store, f, indent=2)


def compute_hash(company_name: str, email: str) -> str:
    """Compute a SHA-256 hash as dedup key."""
    normalized = f"{company_name.strip().lower()}|{email.strip().lower()}"
    return hashlib.sha256(normalized.encode()).hexdigest()


def is_duplicate(lead: Lead) -> bool:
    """Check if this lead has already been contacted."""
    store = _load_store()
    h = compute_hash(lead.company_name, lead.email)
    return h in store


def mark_sent(lead: Lead, original_subject: str = "") -> None:
    """Record that an initial email was sent to this lead."""
    store = _load_store()
    h = compute_hash(lead.company_name, lead.email)
    store[h] = {
        "company": lead.company_name,
        "email": lead.email,
        "decision_maker_name": lead.decision_maker_name,
        "industry": lead.industry,
        "original_subject": original_subject,
        "initial_sent_date": datetime.now().strftime("%Y-%m-%d"),
        "sequence_stage": "initial",
        "followup1_date": None,
        "followup2_date": None,
        "followup3_date": None,
    }
    _save_store()


def mark_followup_sent(company_name: str, email: str, stage: str) -> None:
    """Advance a lead's sequence stage and record the follow-up date."""
    store = _load_store()
    h = compute_hash(company_name, email)
    if h not in store:
        logger.warning("mark_followup_sent: no entry found for %s (%s)", company_name, email)
        return
    entry = store[h]
    entry["sequence_stage"] = stage
    entry[f"{stage}_date"] = datetime.now().strftime("%Y-%m-%d")
    _save_store()


def get_leads_due_for_followup(stage: str, min_days: int, max_days: int) -> list:
    """
    Return store entries where:
    - current sequence_stage == the stage immediately before `stage`
    - days elapsed since that preceding stage was sent are within [min_days, max_days]
    """
    store = _load_store()
    today = datetime.now().date()

    stage_order = ["initial", "followup1", "followup2", "followup3"]
    if stage not in stage_order:
        logger.error("Unknown stage: %s", stage)
        return []

    stage_idx = stage_order.index(stage)
    preceding_stage = stage_order[stage_idx - 1]
    date_field = "initial_sent_date" if preceding_stage == "initial" else f"{preceding_stage}_date"

    results = []
    for h, entry in store.items():
        current_stage = entry.get("sequence_stage", "initial")
        if current_stage != preceding_stage:
            continue
        raw_date = entry.get(date_field)
        if not raw_date:
            continue
        try:
            sent_date = datetime.strptime(raw_date, "%Y-%m-%d").date()
        except ValueError:
            continue
        days_elapsed = (today - sent_date).days
        if min_days <= days_elapsed <= max_days:
            results.append({**entry, "_hash": h})

    return results


def deduplicate_leads(leads: list) -> list:
    """Remove leads already contacted or duplicated within this batch."""
    seen = set()
    unique = []
    for lead in leads:
        h = compute_hash(lead.company_name, lead.email)
        if h not in seen and not is_duplicate(lead):
            seen.add(h)
            unique.append(lead)
    return unique

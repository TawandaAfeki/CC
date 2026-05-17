import logging
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from config import EXCEL_PATH, DATA_DIR
from leads.models import Lead

logger = logging.getLogger(__name__)

# 15-column schema (Phase 2)
HEADERS = [
    "Company Name",           # col 0  (A)
    "Decision-Maker Name",    # col 1  (B)
    "Decision-Maker Title",   # col 2  (C)
    "Email Used",             # col 3  (D)
    "Phone",                  # col 4  (E)
    "LinkedIn URL",           # col 5  (F)
    "Industry",               # col 6  (G)
    "Location",               # col 7  (H)
    "Source",                 # col 8  (I)
    "Company Website",        # col 9  (J)
    "Sequence Stage",         # col 10 (K)
    "Date First Contacted",   # col 11 (L)
    "Date Last Contacted",    # col 12 (M)
    "Next Follow-up Due",     # col 13 (N)
    "Response Status",        # col 14 (O)
]

_LEGACY_COL_COUNT = 5  # old schema had exactly 5 columns


def init_workbook() -> None:
    """
    Create outreach_log.xlsx with the 15-column schema.
    If an old 5-column file exists, rename it to outreach_log_legacy.xlsx
    and create a fresh file with the new schema.
    """
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    if EXCEL_PATH.exists():
        # Peek at the header count to determine schema version
        try:
            wb_check = load_workbook(EXCEL_PATH, read_only=True)
            ws_check = wb_check.active
            header_row = [cell.value for cell in next(ws_check.iter_rows(min_row=1, max_row=1))]
            wb_check.close()
            if len(header_row) < 15:
                # Legacy file — migrate
                legacy_path = EXCEL_PATH.parent / "outreach_log_legacy.xlsx"
                EXCEL_PATH.rename(legacy_path)
                logger.info("Renamed legacy Excel log to %s", legacy_path)
            else:
                # Already new schema — nothing to do
                return
        except Exception as e:
            logger.warning("Could not inspect existing Excel file: %s", e)
            return

    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Outreach Log"
    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
    wb.save(EXCEL_PATH)
    logger.info("Created new 15-column outreach log at %s", EXCEL_PATH)


def log_email(lead: Lead, sent_at: datetime, original_subject: str = "") -> None:
    """Append a 15-column row for a freshly sent initial email."""
    init_workbook()
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    next_followup = (sent_at + timedelta(days=3)).strftime("%Y-%m-%d")

    ws.append([
        lead.company_name,
        lead.decision_maker_name,
        lead.decision_maker_title,
        lead.email,
        lead.phone,
        lead.linkedin_url,
        lead.industry,
        lead.location,
        lead.source,
        lead.website,
        "initial",
        sent_at.strftime("%Y-%m-%d %H:%M:%S"),
        sent_at.strftime("%Y-%m-%d %H:%M:%S"),
        next_followup,
        "pending",
    ])
    wb.save(EXCEL_PATH)
    logger.info("Logged initial email to %s for %s", lead.email, lead.company_name)


def log_followup(
    email: str,
    company: str,
    dm_name: str,
    industry: str,
    stage: str,
    sent_at: datetime,
) -> None:
    """
    Update the existing row for this lead with follow-up info.
    Looks up by company name (col 0) + email (col 3).
    Falls back to appending a new row if not found.
    """
    init_workbook()
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # Calculate next follow-up date based on stage
    next_days = {"followup1": 4, "followup2": 7, "followup3": None}
    days_ahead = next_days.get(stage)
    next_followup = (sent_at + timedelta(days=days_ahead)).strftime("%Y-%m-%d") if days_ahead else "N/A"

    # Search for matching row (skip header)
    updated = False
    for row in ws.iter_rows(min_row=2):
        row_company = row[0].value or ""
        row_email = row[3].value or ""
        if row_company.strip().lower() == company.strip().lower() and row_email.strip().lower() == email.strip().lower():
            row[10].value = stage                             # Sequence Stage
            row[12].value = sent_at.strftime("%Y-%m-%d %H:%M:%S")  # Date Last Contacted
            row[13].value = next_followup                     # Next Follow-up Due
            updated = True
            logger.info("Updated Excel row for %s (%s) → %s", company, email, stage)
            break

    if not updated:
        # Append a new row (lead was logged before schema migration)
        ws.append([
            company,
            dm_name,
            "",
            email,
            "",
            "",
            industry,
            "",
            "",
            "",
            stage,
            "",
            sent_at.strftime("%Y-%m-%d %H:%M:%S"),
            next_followup,
            "pending",
        ])
        logger.info("Appended new row for follow-up %s to %s (%s)", stage, company, email)

    wb.save(EXCEL_PATH)


def get_todays_count() -> int:
    """
    Count initial emails sent today.
    Checks col 10 (Sequence Stage) == 'initial' AND col 11 (Date First Contacted) starts with today.
    Falls back to legacy col 4 check if the file uses old schema.
    """
    if not EXCEL_PATH.exists():
        return 0

    try:
        wb = load_workbook(EXCEL_PATH, read_only=True)
        ws = wb.active
        today_str = datetime.now().strftime("%Y-%m-%d")

        # Detect schema by header count
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        is_new_schema = len(header_row) >= 15

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if is_new_schema:
                stage = row[10] if len(row) > 10 else None
                date_val = row[11] if len(row) > 11 else None
                if stage == "initial" and date_val and str(date_val).startswith(today_str):
                    count += 1
            else:
                # Legacy: col 4 is "Date Time of email sent"
                date_val = row[4] if len(row) > 4 else None
                if date_val and str(date_val).startswith(today_str):
                    count += 1
        wb.close()
        return count
    except Exception as e:
        logger.error("Could not read today's count from Excel: %s", e)
        return 0
